"""
╔══════════════════════════════════════════════════════════════════════════════╗
║                                                                              ║
║   ██████╗██████╗ ██╗   ██╗██████╗ ████████╗ ██████╗      █████╗ ██╗         ║
║  ██╔════╝██╔══██╗╚██╗ ██╔╝██╔══██╗╚══██╔══╝██╔═══██╗    ██╔══██╗██║         ║
║  ██║     ██████╔╝ ╚████╔╝ ██████╔╝   ██║   ██║   ██║    ███████║██║         ║
║  ██║     ██╔══██╗  ╚██╔╝  ██╔═══╝    ██║   ██║   ██║    ██╔══██║██║         ║
║  ╚██████╗██║  ██║   ██║   ██║        ██║   ╚██████╔╝    ██║  ██║███████╗    ║
║   ╚═════╝╚═╝  ╚═╝   ╚═╝   ╚═╝        ╚═╝    ╚═════╝     ╚═╝  ╚═╝╚══════╝    ║
║                                                                              ║
║              Next-Generation Cryptocurrency Intelligence Platform            ║
║              Selenium  ·  pandas  ·  openpyxl  ·  CoinMarketCap             ║
╚══════════════════════════════════════════════════════════════════════════════╝

HOW THIS FILE IS ORGANISED
═══════════════════════════
  SECTION 1  ──  Imports & Setup
  SECTION 2  ──  Configuration (edit this to personalise the tracker)
  SECTION 3  ──  Terminal Color Palette
  SECTION 4  ──  Number Formatters
  SECTION 5  ──  Intelligence Engine  (sentiment · dominance · tiers · alerts)
  SECTION 6  ──  Chrome WebDriver Setup
  SECTION 7  ──  CoinMarketCap Scraper
  SECTION 8  ──  CSV Exporters  (3 separate CSV files)
  SECTION 9  ──  Excel Exporter  (4-sheet workbook)
  SECTION 10 ──  Terminal Dashboard  (banner · table · charts · portfolio)
  SECTION 11 ──  Main Entry Point

OUTPUT FILES  (all saved inside  output/  folder)
═════════════════════════════════════════════════
  crypto_history.csv          ← Appended log of EVERY run  (grows over time)
  crypto_session_latest.csv   ← Current session snapshot   (overwritten each run)
  portfolio_history.csv       ← Portfolio P&L log          (appended every run)
  CryptoPulse_Report.xlsx     ← 4-sheet Excel workbook     (overwritten each run)
  session_report.txt          ← Human-readable text log    (appended every run)
"""


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 1 — IMPORTS & SETUP
# ══════════════════════════════════════════════════════════════════════════════
# Standard library — built into Python, no installation needed
import time          # Controls scraping delays so the page loads fully
import os            # File-system operations: creating folders, checking file existence
import csv           # Reading & writing CSV files row by row
import sys           # Lets us exit the program cleanly with sys.exit()
from datetime import datetime   # Timestamps on every record

# Third-party libraries — installed via:  pip install -r requirements.txt
import pandas as pd                          # Data analysis on historical CSV data

from selenium import webdriver               # Controls Chrome browser programmatically
from selenium.webdriver.chrome.service import Service   # Manages ChromeDriver process
from selenium.webdriver.chrome.options import Options   # Chrome launch flags (headless etc.)
from selenium.webdriver.common.by import By             # How to locate HTML elements
from selenium.webdriver.support.ui import WebDriverWait # Wait until page is ready
from selenium.webdriver.support import expected_conditions as EC  # Conditions to wait for
from webdriver_manager.chrome import ChromeDriverManager  # Auto-downloads correct ChromeDriver


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 2 — CONFIGURATION
#  ✏️  Edit values in this block to personalise the tracker.
# ══════════════════════════════════════════════════════════════════════════════

# All output files are saved inside this subfolder.
# It is created automatically if it doesn't exist.
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CONFIG = {
    # ── Scraper settings ──────────────────────────────────────────────────────
    "headless" : True,   # True  → Chrome runs silently in background (recommended)
                         # False → Chrome window opens visibly (good for debugging)
    "top_n"    : 10,     # How many top coins to scrape from CoinMarketCap

    # ── Output file paths ────────────────────────────────────────────────────
    # All paths are inside the OUTPUT_DIR folder defined above.
    "csv_history"    : os.path.join(OUTPUT_DIR, "crypto_history.csv"),
    "csv_session"    : os.path.join(OUTPUT_DIR, "crypto_session_latest.csv"),
    "csv_portfolio"  : os.path.join(OUTPUT_DIR, "portfolio_history.csv"),
    "excel_file"     : os.path.join(OUTPUT_DIR, "CryptoPulse_Report.xlsx"),
    "report_file"    : os.path.join(OUTPUT_DIR, "session_report.txt"),

    # ── Alert thresholds ─────────────────────────────────────────────────────
    "alert_surge"  : 5.0,    # Fire a SURGE alert if a coin gains more than this %
    "alert_crash"  : -5.0,   # Fire a CRASH alert if a coin drops more than this %

    # ── Dashboard ────────────────────────────────────────────────────────────
    "top_gainers" : 3,   # How many top gainers to show in the movers panel
    "top_losers"  : 3,   # How many top losers  to show in the movers panel

    # ── Optional price filters (set to None to disable) ──────────────────────
    "min_price" : None,   # e.g.  1.0  → only show coins priced above $1
    "max_price" : None,   # e.g.  500.0 → only show coins priced below $500

    # ── Your portfolio  (coin symbol → quantity you own) ─────────────────────
    # Change these values to match what you actually hold.
    "portfolio" : {
        "BTC"  : 0.01,    #  0.01 Bitcoin
        "ETH"  : 0.50,    #  0.50 Ethereum
        "BNB"  : 2.00,    #  2    BNB
        "SOL"  : 5.00,    #  5    Solana
    },
}

# CoinMarketCap homepage — the page we scrape
CMC_URL = "https://coinmarketcap.com/"


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 3 — TERMINAL COLOR PALETTE
#  ANSI escape codes that colorize text in the terminal.
# ══════════════════════════════════════════════════════════════════════════════
class C:
    RESET   = "\033[0m"    # Resets all formatting
    BOLD    = "\033[1m"    # Bold text
    DIM     = "\033[2m"    # Dimmed / faint text
    GREEN   = "\033[92m"   # Bright green  (gains, success)
    RED     = "\033[91m"   # Bright red    (losses, errors)
    YELLOW  = "\033[93m"   # Bright yellow (warnings, headers)
    CYAN    = "\033[96m"   # Bright cyan   (symbols, highlights)
    MAGENTA = "\033[95m"   # Magenta       (market cap, portfolio box)
    WHITE   = "\033[97m"   # Bright white  (coin names)
    BLUE    = "\033[94m"   # Bright blue   (chart labels)
    ORANGE  = "\033[33m"   # Orange        (MID tier)

def col(text, color):
    """Wrap text in a terminal color code, then reset."""
    return "{}{}{}".format(color, text, C.RESET)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 4 — NUMBER FORMATTERS
#  Helper functions that turn raw floats into readable strings.
#  Keeping formatting in functions avoids nested-quote errors in f-strings.
# ══════════════════════════════════════════════════════════════════════════════

def fmt_price(p):
    """
    Format a price smartly based on its magnitude:
      ≥ $1000  →  $62,123.45   (2 decimal places, comma-separated)
      ≥ $1     →  $1.6509      (4 decimal places for mid-range coins)
      < $1     →  $0.000312    (6 decimal places for micro-cap coins)
    """
    if p >= 1000: return "${:,.2f}".format(p)
    if p >= 1:    return "${:,.4f}".format(p)
    return "${:.6f}".format(p)

def fmt_change(c):
    """Format 24h change with a leading + for gains, e.g. +2.45% or -1.23%"""
    return ("+{:.2f}%".format(c)) if c >= 0 else ("{:.2f}%".format(c))

def fmt_mcap(m):
    """
    Format market cap into human-readable shorthand:
      Trillions → $1.22T
      Billions  → $197.86B
      Millions  → $12.84M
    """
    if m >= 1e12: return "${:.2f}T".format(m / 1e12)
    if m >= 1e9:  return "${:.2f}B".format(m / 1e9)
    if m >= 1e6:  return "${:.2f}M".format(m / 1e6)
    return "${:,.0f}".format(m) if m > 0 else "N/A"

def fmt_vol(v):
    """Volume uses the same scale as market cap."""
    return fmt_mcap(v)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 5 — INTELLIGENCE ENGINE
#  These functions compute insights from the raw scraped data.
# ══════════════════════════════════════════════════════════════════════════════

def classify_tier(market_cap):
    """
    Categorise a coin into a size tier based on market cap.
    This mirrors how professional analysts classify assets:
      MEGA  ≥ $200B  — Bitcoin, Ethereum  (blue chip)
      LARGE ≥ $10B   — BNB, XRP, SOL      (established)
      MID   ≥ $1B    — Smaller altcoins   (speculative)
      SMALL < $1B    — Very small coins   (high risk)
    Returns (label_string, terminal_color).
    """
    if market_cap >= 200e9: return ("MEGA",  C.CYAN)
    if market_cap >= 10e9:  return ("LARGE", C.GREEN)
    if market_cap >= 1e9:   return ("MID",   C.ORANGE)
    return ("SMALL", C.RED)


def compute_sentiment(coins):
    """
    Calculate a Market Sentiment Score from 0 to 100.

    Formula:
      60% weight → what fraction of coins are gaining (gainer ratio)
      40% weight → average % change scaled to a 0-1 range assuming ±10% is extreme

    Labels:
      75-100 → EXTREME GREED  (market is very bullish, potential bubble)
      60-74  → GREED           (more buyers than sellers)
      45-59  → NEUTRAL         (balanced market)
      30-44  → FEAR            (more sellers, caution advised)
       0-29  → EXTREME FEAR    (panic selling, potential buy opportunity)

    Returns (score_int, colored_label_string).
    """
    if not coins:
        return 50, "NEUTRAL"

    gainers       = [c for c in coins if c["change_24h"] > 0]
    avg_change    = sum(c["change_24h"] for c in coins) / len(coins)
    gainer_ratio  = len(gainers) / len(coins)
    change_score  = min(max((avg_change + 10) / 20, 0), 1)  # Map [-10, +10] → [0, 1]
    score         = int((gainer_ratio * 0.6 + change_score * 0.4) * 100)

    if score >= 75:   label = col("EXTREME GREED  🔥", C.RED)
    elif score >= 60: label = col("GREED  📈",          C.ORANGE)
    elif score >= 45: label = col("NEUTRAL  ➡️",        C.YELLOW)
    elif score >= 30: label = col("FEAR  📉",           C.CYAN)
    else:             label = col("EXTREME FEAR  ❄️",  C.BLUE)

    return score, label


def compute_dominance(coins):
    """
    Calculate each coin's share of the total market cap within the top-N set.

    Example: If BTC has $1.2T of the $2T combined top-10 cap → BTC dominance = 60%.

    Returns a dict: { "BTC": 60.0, "ETH": 10.2, ... }
    """
    total = sum(c["market_cap"] for c in coins if c["market_cap"] > 0)
    if total == 0:
        return {}
    return {
        c["symbol"]: (c["market_cap"] / total) * 100
        for c in coins if c["market_cap"] > 0
    }


def generate_alerts(coins):
    """
    Scan coins for notable price movements and return alert messages.

    Three alert types:
      🚀 SURGE   — coin gained more than alert_surge threshold in 24h
      🔻 CRASH   — coin dropped more than alert_crash threshold in 24h
      ⚡ VOLATILE — the single biggest mover (positive or negative) if it moved > 3%

    Returns a list of pre-colored strings, or an empty list if quiet market.
    """
    alerts = []
    for c in coins:
        ch, sym = c["change_24h"], c["symbol"]
        if ch >= CONFIG["alert_surge"]:
            alerts.append(col(
                "  🚀  SURGE   : {} soared {} in 24h!".format(sym, fmt_change(ch)),
                C.GREEN
            ))
        elif ch <= CONFIG["alert_crash"]:
            alerts.append(col(
                "  🔻  CRASH   : {} plunged {} in 24h!".format(sym, fmt_change(ch)),
                C.RED
            ))

    # Biggest mover alert (regardless of direction)
    top = max(coins, key=lambda x: abs(x["change_24h"]))
    if abs(top["change_24h"]) >= 3.0:
        alerts.append(col(
            "  ⚡  VOLATILE: {} is biggest mover at {}".format(
                top["symbol"], fmt_change(top["change_24h"])),
            C.YELLOW
        ))

    return alerts


def apply_filters(coins):
    """
    Remove coins that fall outside the optional price range in CONFIG.
    If both min_price and max_price are None, all coins pass through unchanged.
    """
    result = coins[:]
    if CONFIG["min_price"] is not None:
        result = [c for c in result if c["price"] >= CONFIG["min_price"]]
    if CONFIG["max_price"] is not None:
        result = [c for c in result if c["price"] <= CONFIG["max_price"]]
    return result


def calc_portfolio(coins):
    """
    Cross-reference the user's holdings (CONFIG["portfolio"]) with live prices.
    For each holding where we have a price, compute: value_usd = price × quantity.
    Returns a list of dicts, one per coin held.
    """
    price_map = {c["symbol"]: c["price"] for c in coins}
    return [
        {
            "symbol"    : sym,
            "qty"       : qty,
            "price"     : price_map[sym],
            "value_usd" : price_map[sym] * qty,
        }
        for sym, qty in CONFIG["portfolio"].items()
        if sym in price_map
    ]


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 6 — CHROME WEBDRIVER SETUP
# ══════════════════════════════════════════════════════════════════════════════

def create_driver():
    """
    Build and return a configured Chrome WebDriver.

    Key settings:
      --headless=new        → Run Chrome without a visible window
      --no-sandbox          → Required inside virtual environments / Docker
      --disable-dev-shm-usage → Prevents crashes on low-memory systems
      --disable-blink-features=AutomationControlled → Hides the fact that
                              Chrome is being automated (anti-bot bypass)
      user-agent            → Sets a realistic browser identity string
      excludeSwitches       → Removes the "Chrome is being controlled" banner
      CDP command           → Patches navigator.webdriver to undefined so
                              JavaScript on the page cannot detect Selenium
    """
    opts = Options()

    if CONFIG["headless"]:
        opts.add_argument("--headless=new")

    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--log-level=3")   # Suppress Chrome's verbose console output
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    opts.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    opts.add_experimental_option("useAutomationExtension", False)

    # webdriver_manager automatically downloads the ChromeDriver version that
    # matches your installed Chrome — no manual driver management needed.
    svc    = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=svc, options=opts)

    # Patch the webdriver property so anti-bot scripts see it as undefined
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"}
    )
    return driver


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 7 — COINMARKETCAP SCRAPER
# ══════════════════════════════════════════════════════════════════════════════

def scrape(driver):
    """
    Navigate to CoinMarketCap and extract the top-N cryptocurrency rows.

    HOW IT WORKS:
      1. driver.get(URL)       → Chrome opens the CoinMarketCap homepage
      2. WebDriverWait         → We pause until the <table> rows appear in the DOM
                                 (the page uses JavaScript to render, so we must wait)
      3. time.sleep(3)         → Extra buffer for dynamic values to fully populate
      4. find_elements         → Grab all <tr> rows from the main table body
      5. Loop through top_n    → Parse each row's <td> cells into a Python dict

    CELL LAYOUT  (approximate — CoinMarketCap updates their HTML periodically):
      cells[1] → Rank number
      cells[2] → Name + Symbol  (two lines of text in one cell)
      cells[3] → Price
      cells[5] → 24h change %
      cells[7] → Market cap
      cells[8] → Volume 24h  (may be absent on some rows)

    Returns a list of dicts, each representing one coin.
    """
    print(col("  ⚡  Connecting to CoinMarketCap...", C.CYAN))
    driver.get(CMC_URL)

    try:
        # Wait up to 20 seconds for the first table row to appear
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
        )
        time.sleep(3)   # Give JS time to fill all price cells
    except Exception:
        print(col("  ✗  Page load timed out — check internet connection.", C.RED))
        return []

    rows  = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
    coins = []
    ts    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row in rows[:CONFIG["top_n"]]:
        try:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) < 8:
                continue   # Skip rows that don't have enough columns

            # ── Rank ────────────────────────────────────────────────────────
            rank = cells[1].text.strip()

            # ── Name & Symbol  (same cell, split by newline) ────────────────
            parts  = cells[2].text.strip().split("\n")
            name   = parts[0] if parts else "Unknown"
            symbol = (parts[1] if len(parts) > 1 else "N/A").upper()

            def clean(raw):
                """Strip currency symbols, commas, percent, plus signs."""
                return raw.replace("$","").replace(",","").replace("+","").replace("%","").strip()

            # ── Price ────────────────────────────────────────────────────────
            try:
                price = float(clean(cells[3].text))
            except (ValueError, IndexError):
                price = 0.0

            # ── 24h Change ──────────────────────────────────────────────────
            try:
                change = float(clean(cells[5].text))
            except (ValueError, IndexError):
                change = 0.0

            # ── Market Cap  (CoinMarketCap shows e.g. "$1.22T", "$197B") ────
            try:
                mc_raw = cells[7].text.strip().replace("$","").replace(",","")
                if   mc_raw.endswith("T"): mcap = float(mc_raw[:-1]) * 1e12
                elif mc_raw.endswith("B"): mcap = float(mc_raw[:-1]) * 1e9
                elif mc_raw.endswith("M"): mcap = float(mc_raw[:-1]) * 1e6
                else: mcap = float(mc_raw) if mc_raw not in ("","--","N/A") else 0.0
            except (ValueError, IndexError):
                mcap = 0.0

            # ── Volume 24h  (same format as market cap) ─────────────────────
            try:
                vl_raw = cells[8].text.strip().replace("$","").replace(",","")
                if   vl_raw.endswith("T"): vol = float(vl_raw[:-1]) * 1e12
                elif vl_raw.endswith("B"): vol = float(vl_raw[:-1]) * 1e9
                elif vl_raw.endswith("M"): vol = float(vl_raw[:-1]) * 1e6
                else: vol = float(vl_raw) if vl_raw not in ("","--","N/A") else 0.0
            except (ValueError, IndexError):
                vol = 0.0

            coins.append({
                "rank"       : rank,
                "name"       : name,
                "symbol"     : symbol,
                "price"      : price,
                "change_24h" : change,
                "market_cap" : mcap,
                "volume_24h" : vol,
                "timestamp"  : ts,
            })

        except Exception:
            continue   # Skip any row that throws an unexpected error

    return coins


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 8 — CSV EXPORTERS
#  Three separate CSV files are produced, each serving a different purpose.
# ══════════════════════════════════════════════════════════════════════════════

# ── CSV FILE 1: Historical Log ────────────────────────────────────────────────
def save_csv_history(coins):
    """
    PURPOSE : Long-term historical record of every scrape ever run.
    BEHAVIOR: Appends new rows — never overwrites old data.
    USE FOR : Trend analysis, charting price changes over weeks/months,
              feeding into a visualization dashboard.

    Columns: timestamp · rank · name · symbol · price · change_24h
             market_cap · volume_24h

    SAFETY:
      • Only writes rows where symbol is a short uppercase string (real coins)
        This prevents section headers like "=== SECTION ===" from contaminating
        the history file if the file is ever accidentally merged.
      • The header is written only once (when the file is new).
    """
    path   = CONFIG["csv_history"]
    exists = os.path.isfile(path)
    fields = ["timestamp","rank","name","symbol","price",
              "change_24h","market_cap","volume_24h"]

    with open(path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        if not exists:
            writer.writeheader()   # Write column names only the very first time
        for c in coins:
            # Guard: only write genuine coin rows
            sym = str(c.get("symbol", ""))
            if not sym or sym.startswith("=") or sym.startswith("#") or len(sym) > 12:
                continue
            writer.writerow({k: c.get(k, "") for k in fields})

    print(col("  ✔  History CSV    → " + path, C.GREEN))


# ── CSV FILE 2: Session Snapshot ──────────────────────────────────────────────
def save_csv_session(coins, pnl, dominance, score):
    """
    PURPOSE : A clean, self-contained snapshot of the current session.
    BEHAVIOR: Overwrites the file on every run — always shows latest data.
    USE FOR : Opening in Excel for a quick one-off look, sharing with others,
              importing into Google Sheets.

    The file has THREE sections separated by blank rows:
      ┌─────────────────────────────────────────────────────┐
      │  # CRYPTOPULSE — Session Snapshot                   │
      │  # Timestamp: 2026-06-11 15:02:00                   │
      │  # Sentiment: 76/100                                │
      │                                                     │
      │  === SECTION 1: LIVE MARKET DATA ===                │
      │  Rank, Name, Symbol, Tier, Price, 24H Change, ...   │
      │  1, Bitcoin, BTC, MEGA, 62124.02, +2.45%, ...       │
      │  ...                                                │
      │                                                     │
      │  === SECTION 2: PORTFOLIO SNAPSHOT ===              │
      │  Symbol, Quantity, Price, Value, % of Portfolio     │
      │  BTC, 0.01, 62124.02, 621.24, 21.1%                 │
      │  TOTAL, , , 2946.66, 100%                           │
      │                                                     │
      │  === SECTION 3: MARKET DOMINANCE ===                │
      │  Symbol, Market Cap, Dominance %, Visual Bar        │
      │  BTC, 1220000000000, 64.40%, ████████████████...    │
      └─────────────────────────────────────────────────────┘
    """
    path = CONFIG["csv_session"]
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)

        # ── File header (metadata rows, prefixed with # so tools can skip them)
        w.writerow(["# CRYPTOPULSE — Live Session Snapshot"])
        w.writerow(["# Timestamp", ts])
        w.writerow(["# Sentiment Score", "{}/100".format(score)])
        w.writerow(["# Coins Tracked", str(len(coins))])
        w.writerow([])   # Blank separator row

        # ── SECTION 1: Market Data ───────────────────────────────────────────
        w.writerow(["=== SECTION 1: LIVE MARKET DATA ==="])
        w.writerow(["Rank","Name","Symbol","Tier",
                    "Price (USD)","24H Change (%)","Market Cap (USD)",
                    "Volume 24H (USD)","Timestamp"])
        for c in coins:
            tier_lbl, _ = classify_tier(c["market_cap"])
            w.writerow([
                c["rank"],
                c["name"],
                c["symbol"],
                tier_lbl,
                "{:.6f}".format(c["price"]),        # Full precision price
                "{:+.2f}".format(c["change_24h"]),  # e.g. +2.45 or -1.23
                "{:.0f}".format(c["market_cap"]),
                "{:.0f}".format(c["volume_24h"]),
                c["timestamp"],
            ])
        w.writerow([])

        # ── SECTION 2: Portfolio Snapshot ────────────────────────────────────
        w.writerow(["=== SECTION 2: PORTFOLIO SNAPSHOT ==="])
        w.writerow(["Symbol","Quantity","Price (USD)","Value (USD)","% of Portfolio"])
        total_val = sum(r["value_usd"] for r in pnl) if pnl else 0
        for r in pnl:
            pct = (r["value_usd"] / total_val * 100) if total_val > 0 else 0
            w.writerow([
                r["symbol"],
                "{:.8f}".format(r["qty"]),       # Full precision quantity
                "{:.6f}".format(r["price"]),
                "{:.2f}".format(r["value_usd"]),
                "{:.2f}%".format(pct),
            ])
        if pnl:
            w.writerow(["TOTAL", "", "", "{:.2f}".format(total_val), "100.00%"])
        w.writerow([])

        # ── SECTION 3: Market Dominance ──────────────────────────────────────
        w.writerow(["=== SECTION 3: MARKET DOMINANCE MAP ==="])
        w.writerow(["Symbol","Market Cap (USD)","Dominance (%)","Visual Bar"])
        for sym, pct in sorted(dominance.items(), key=lambda x: -x[1]):
            mc  = next((c["market_cap"] for c in coins if c["symbol"] == sym), 0)
            bar = "█" * max(1, int(pct / 2))   # 2% per block
            w.writerow([sym, "{:.0f}".format(mc), "{:.2f}%".format(pct), bar])

    print(col("  ✔  Session CSV    → " + path, C.GREEN))


# ── CSV FILE 3: Portfolio History Log ─────────────────────────────────────────
def save_csv_portfolio(pnl):
    """
    PURPOSE : Track how your portfolio's value changes over multiple sessions.
    BEHAVIOR: Appends timestamped rows — never overwrites.
    USE FOR : Plotting portfolio performance over time, building a personal
              investment journal, tracking allocation drift.

    Each run adds one row per held coin:
      timestamp · symbol · quantity · price_usd · value_usd · portfolio_pct
    """
    path   = CONFIG["csv_portfolio"]
    exists = os.path.isfile(path)
    ts     = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total  = sum(r["value_usd"] for r in pnl) if pnl else 0

    with open(path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["timestamp","symbol","quantity",
                        "price_usd","value_usd","portfolio_pct"]
        )
        if not exists:
            writer.writeheader()
        for r in pnl:
            pct = (r["value_usd"] / total * 100) if total else 0
            writer.writerow({
                "timestamp"     : ts,
                "symbol"        : r["symbol"],
                "quantity"      : "{:.8f}".format(r["qty"]),
                "price_usd"     : "{:.6f}".format(r["price"]),
                "value_usd"     : "{:.2f}".format(r["value_usd"]),
                "portfolio_pct" : "{:.2f}%".format(pct),
            })

    print(col("  ✔  Portfolio CSV  → " + path, C.GREEN))


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 9 — EXCEL EXPORTER  (4-sheet professional workbook)
# ══════════════════════════════════════════════════════════════════════════════

def save_excel(coins, pnl, dominance, score, label):
    """
    Build and save a richly-formatted Excel workbook with four sheets:

    Sheet 1 — 📊 Live Market
        Full market table with Tier badges (colour-coded cells), 24h change
        in green/red, market cap, volume, plus auto-formula totals.

    Sheet 2 — 💼 Portfolio
        Your holdings with live prices, USD values, portfolio % share,
        a =SUM() total row, and a pie chart of allocation.

    Sheet 3 — 🏆 Dominance Map
        Each coin's share of combined market cap as a % and a visual
        bar made of █ characters, plus a horizontal bar chart.

    Sheet 4 — 📈 History & Trends
        Every row ever appended to crypto_history.csv, with change cells
        colour-coded green/red.

    STYLE GUIDE used throughout:
      Dark header rows  (hex #0D1117)  — near-black background
      Gold header text  (hex #F0B90B)  — Binance-style gold
      Green cells       (hex #0ECB81)  — positive 24h change ≥ +1%
      Red   cells       (hex #F6465D)  — negative 24h change ≤ -1%
      Alternating rows  white / #F6F8FA for readability
      Freeze panes      — header row stays visible when scrolling
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference, PieChart
    except ImportError:
        print(col("  ⚠  openpyxl not installed. Run: pip install openpyxl", C.YELLOW))
        return

    # ── Hex color palette ─────────────────────────────────────────────────────
    DARK  = "0D1117";  DARK2 = "161B22";  GOLD  = "F0B90B"
    GRN   = "0ECB81";  RED   = "F6465D";  BLU   = "1A73E8"
    WHT   = "FFFFFF";  LGR   = "F6F8FA";  MGR   = "E1E4E8"
    ORG   = "FF6B35"

    ts_now = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")

    # ── Mini style helpers (avoids repeating the same 4-5 lines everywhere) ──
    def sf(h):
        return PatternFill("solid", fgColor=h)

    def th():
        s = Side(style="thin", color=MGR)
        return Border(left=s, right=s, top=s, bottom=s)

    def tk():
        s = Side(style="medium", color="888888")
        return Border(left=s, right=s, top=s, bottom=s)

    def F(sz=9, bold=False, color="1C1C1C"):
        return Font(name="Arial", size=sz, bold=bold, color=color)

    def ctr(): return Alignment(horizontal="center", vertical="center")
    def lft(): return Alignment(horizontal="left",   vertical="center", indent=1)
    def rgt(): return Alignment(horizontal="right",  vertical="center")

    # ── Reusable row builders ─────────────────────────────────────────────────
    def title(ws, row, text, end_col, bg=DARK, fc=GOLD, sz=12):
        """Dark full-width title row spanning all columns."""
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=end_col)
        c = ws.cell(row, 1, "   " + text)
        c.font = Font(name="Arial", size=sz, bold=True, color=fc)
        c.fill = sf(bg); c.alignment = lft()
        ws.row_dimensions[row].height = 32

    def subtitle(ws, row, text, end_col, bg=DARK2):
        """Dimmed italic subtitle row beneath the title."""
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=end_col)
        c = ws.cell(row, 1, "   " + text)
        c.font = Font(name="Arial", size=9, italic=True, color="8B949E")
        c.fill = sf(bg); c.alignment = lft()
        ws.row_dimensions[row].height = 18

    def spacer(ws, row, end_col):
        """Thin dark spacer row between title block and headers."""
        for ci in range(1, end_col + 1):
            ws.cell(row, ci).fill = sf(DARK)
        ws.row_dimensions[row].height = 5

    def headers(ws, row, lbls, wids, bg=DARK, fc=GOLD):
        """Bold header row with gold text on dark background."""
        for i, lbl in enumerate(lbls, 1):
            c = ws.cell(row, i, lbl)
            c.font = Font(name="Arial", size=9, bold=True, color=fc)
            c.fill = sf(bg); c.alignment = ctr(); c.border = th()
        for i, w in enumerate(wids, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[row].height = 22

    def cell(ws, r, ci, val, aln="right", bold=False,
             fc="1C1C1C", bg=None, fmt=None):
        """Write a single data cell with consistent styling."""
        c = ws.cell(r, ci, val)
        c.font = F(9, bold, fc)
        c.alignment = ctr() if aln == "center" else (rgt() if aln == "right" else lft())
        c.fill   = sf(bg if bg else WHT)
        c.border = th()
        if fmt:
            c.number_format = fmt
        return c

    def change_cell(ws, r, ci, ch, bg=WHT):
        """
        Special cell for 24h % change — auto-coloured:
          Green fill + white bold text if change ≥ +1%
          Red   fill + white bold text if change ≤ -1%
          Subtle green/red text only   if between -1% and +1%
        """
        c = ws.cell(r, ci, ch / 100)   # Store as decimal (Excel formats as %)
        c.number_format = "+0.00%;-0.00%"
        c.alignment = ctr(); c.border = th()
        if ch >= 1.0:
            c.fill = sf(GRN); c.font = Font(name="Arial", size=9, bold=True, color=WHT)
        elif ch <= -1.0:
            c.fill = sf(RED); c.font = Font(name="Arial", size=9, bold=True, color=WHT)
        else:
            c.fill = sf(bg)
            c.font = Font(name="Arial", size=9,
                          color="006400" if ch >= 0 else "8B0000")

    def footer(ws, row, label_end_col, label, val_col, formula, fmt="$#,##0"):
        """
        Two-part footer row:
          Cols 1 → label_end_col : merged dark label cell
          Col val_col            : formula/value cell (outside the merge)
        The value column is intentionally placed OUTSIDE the merge to
        avoid the MergedCell read-only error.
        """
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=label_end_col)
        lc = ws.cell(row, 1, "   " + label)
        lc.font = Font(name="Arial", size=10, bold=True, color=GOLD)
        lc.fill = sf(DARK); lc.alignment = lft(); lc.border = tk()
        # Dark fill on the interior merged cols (they're read-only MergedCell
        # objects, so we use a try/except)
        for ci in range(2, label_end_col + 1):
            try:
                ws.cell(row, ci).fill   = sf(DARK)
                ws.cell(row, ci).border = tk()
            except AttributeError:
                pass
        # Value cell — always outside the merge range
        vc = ws.cell(row, val_col, formula)
        vc.number_format = fmt
        vc.font = Font(name="Arial", size=11, bold=True, color=GOLD)
        vc.fill = sf(DARK); vc.alignment = ctr(); vc.border = tk()
        ws.row_dimensions[row].height = 26

    def tier_badge(ws, r, ci, mc):
        """Write a Tier badge cell with background matching the tier."""
        lbl, _ = classify_tier(mc)
        bg_map = {"MEGA": BLU, "LARGE": GRN, "MID": ORG, "SMALL": RED}
        c = ws.cell(r, ci, lbl)
        c.font = Font(name="Arial", size=9, bold=True, color=WHT)
        c.fill = sf(bg_map.get(lbl, DARK))
        c.alignment = ctr(); c.border = th()

    # ═════════════════════════════════════════════════════════════════════════
    wb = Workbook()

    # ══════════════════════════════════════
    #  SHEET 1 — LIVE MARKET
    # ══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📊 Live Market"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A5"   # Rows 1-4 (title + headers) stay fixed when scrolling

    COLS1 = 9
    title(ws1, 1,
          "CRYPTOPULSE  ·  Live Market Intelligence  ·  " + ts_now,
          COLS1, sz=12)
    subtitle(ws1, 2,
             "Sentiment: {}/100  |  Gainers: {}  |  Losers: {}  |  Avg 24H: {:+.2f}%".format(
                 score,
                 len([c for c in coins if c["change_24h"] > 0]),
                 len([c for c in coins if c["change_24h"] < 0]),
                 sum(c["change_24h"] for c in coins) / len(coins) if coins else 0,
             ), COLS1, bg=DARK2)
    subtitle(ws1, 3,
             "Source: CoinMarketCap  |  Selenium WebDriver  |  Top {} Coins".format(len(coins)),
             COLS1)
    spacer(ws1, 4, COLS1)

    headers(ws1, 5,
            ["#","Name","Symbol","Tier","Price (USD)","24H Change",
             "Market Cap","Volume 24H","Timestamp"],
            [5, 17, 9, 8, 15, 14, 20, 18, 22])

    for ri, c in enumerate(coins, start=6):
        bg = LGR if ri % 2 == 0 else WHT
        cell(ws1, ri, 1, c["rank"],       "center", bg=bg)
        cell(ws1, ri, 2, c["name"],       "left",   bold=True, bg=bg)
        cell(ws1, ri, 3, c["symbol"],     "center", bold=True, fc=BLU, bg=bg)
        tier_badge(ws1, ri, 4, c["market_cap"])
        cell(ws1, ri, 5, c["price"],      "right",  bg=bg, fmt="$#,##0.0000")
        change_cell(ws1, ri, 6, c["change_24h"], bg)
        cell(ws1, ri, 7, c["market_cap"], "right",  bg=bg, fmt="$#,##0")
        cell(ws1, ri, 8, c["volume_24h"], "right",  bg=bg, fmt="$#,##0")
        cell(ws1, ri, 9, c["timestamp"],  "center", fc="888888", bg=bg)
        ws1.row_dimensions[ri].height = 20

    last1 = 5 + len(coins)
    footer(ws1, last1 + 2, 4,
           "TOTAL MARKET CAP  (Top {})".format(len(coins)),
           5, "=SUM(G6:G{})".format(last1))
    footer(ws1, last1 + 3, 4,
           "AVERAGE 24H CHANGE  (Top {})".format(len(coins)),
           5, "=AVERAGE(F6:F{})".format(last1), fmt="+0.00%;-0.00%")

    # ══════════════════════════════════════
    #  SHEET 2 — PORTFOLIO
    # ══════════════════════════════════════
    ws2 = wb.create_sheet("💼 Portfolio")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A5"

    COLS2 = 6
    title(ws2, 1, "PORTFOLIO SNAPSHOT  ·  Real-Time P&L  ·  " + ts_now, COLS2)
    subtitle(ws2, 2,
             "Holdings valued at live prices  |  Edit CONFIG['portfolio'] to add your coins",
             COLS2)
    spacer(ws2, 3, COLS2)
    headers(ws2, 4,
            ["Coin","Symbol","Quantity Held","Live Price (USD)","Value (USD)","% of Portfolio"],
            [16, 10, 16, 18, 16, 17])

    total_val = sum(r["value_usd"] for r in pnl) if pnl else 0
    for ri, r in enumerate(pnl, start=5):
        bg  = LGR if ri % 2 == 0 else WHT
        pct = (r["value_usd"] / total_val * 100) if total_val else 0
        cell(ws2, ri, 1, r["symbol"],    "left",   bold=True, bg=bg)
        cell(ws2, ri, 2, r["symbol"],    "center", bold=True, fc=BLU, bg=bg)
        cell(ws2, ri, 3, r["qty"],       "right",  bg=bg, fmt="0.0000####")
        cell(ws2, ri, 4, r["price"],     "right",  bg=bg, fmt="$#,##0.0000")
        cell(ws2, ri, 5, r["value_usd"], "right",  bold=True, fc="006400",
             bg=bg, fmt="$#,##0.00")
        cell(ws2, ri, 6, pct / 100,      "center", bg=bg, fmt="0.00%")
        ws2.row_dimensions[ri].height = 20

    tr = 5 + len(pnl)   # Row immediately after last data row
    # Footer: label in cols 1-5 (merged), value in col 6 (outside merge)
    footer(ws2, tr + 1, 5, "TOTAL PORTFOLIO VALUE",
           6, "=SUM(E5:E{})".format(tr), fmt="$#,##0.00")

    # Pie chart — portfolio allocation
    if len(pnl) >= 2:
        pie = PieChart()
        pie.title  = "Portfolio Allocation"
        pie.style  = 26
        pie.height = 13
        pie.width  = 17
        pie.add_data(
            Reference(ws2, min_col=5, min_row=4, max_row=4 + len(pnl)),
            titles_from_data=True
        )
        pie.set_categories(
            Reference(ws2, min_col=1, min_row=5, max_row=4 + len(pnl))
        )
        ws2.add_chart(pie, "A{}".format(tr + 4))

    # ══════════════════════════════════════
    #  SHEET 3 — DOMINANCE MAP
    # ══════════════════════════════════════
    ws3 = wb.create_sheet("🏆 Dominance Map")
    ws3.sheet_view.showGridLines = False

    COLS3 = 4
    title(ws3, 1,
          "MARKET DOMINANCE  ·  Share Within Top-{} Coins".format(len(coins)),
          COLS3)
    subtitle(ws3, 2,
             "Each coin's % of combined market cap  |  Larger bar = greater dominance",
             COLS3)
    spacer(ws3, 3, COLS3)
    headers(ws3, 4,
            ["Symbol","Market Cap (USD)","Dominance %","Visual Dominance Bar"],
            [10, 22, 14, 40])

    sorted_dom = sorted(dominance.items(), key=lambda x: -x[1])
    for ri, (sym, pct) in enumerate(sorted_dom, start=5):
        bg  = LGR if ri % 2 == 0 else WHT
        bar = ("█" * max(1, int(pct / 2))).ljust(34)
        mc  = next((c["market_cap"] for c in coins if c["symbol"] == sym), 0)
        cell(ws3, ri, 1, sym,     "center", bold=True, fc=BLU, bg=bg)
        cell(ws3, ri, 2, mc,      "right",  bg=bg, fmt="$#,##0")
        cell(ws3, ri, 3, pct/100, "center", bg=bg, fmt="0.00%")
        bc = ws3.cell(ri, 4, bar)
        bc.font = Font(
            name="Courier New", size=9,
            color=GRN if pct >= 20 else (ORG if pct >= 5 else "BBBBBB")
        )
        bc.fill = sf(bg); bc.border = th()
        bc.alignment = lft()
        ws3.row_dimensions[ri].height = 20

    # Horizontal bar chart
    n = len(sorted_dom)
    bch = BarChart()
    bch.type   = "bar"; bch.style = 10
    bch.title  = "Market Dominance (%)"; bch.height = 14; bch.width = 18
    bch.add_data(Reference(ws3, min_col=3, min_row=4, max_row=4+n),
                 titles_from_data=True)
    bch.set_categories(Reference(ws3, min_col=1, min_row=5, max_row=4+n))
    ws3.add_chart(bch, "A{}".format(7 + n))

    # ══════════════════════════════════════
    #  SHEET 4 — HISTORY & TRENDS
    # ══════════════════════════════════════
    ws4 = wb.create_sheet("📈 History & Trends")
    ws4.sheet_view.showGridLines = False

    COLS4 = 8
    title(ws4, 1, "HISTORICAL PRICE LOG  ·  All Scrape Sessions", COLS4)
    subtitle(ws4, 2,
             "Appended every run  |  Use for trend analysis, backtesting & forecasting",
             COLS4)
    spacer(ws4, 3, COLS4)
    headers(ws4, 4,
            ["Timestamp","#","Name","Symbol",
             "Price (USD)","24H Change","Market Cap","Volume 24H"],
            [22, 5, 17, 9, 16, 14, 20, 16])

    hist_path = CONFIG["csv_history"]
    if os.path.isfile(hist_path):
        df = pd.read_csv(hist_path)
        for ri, (_, row) in enumerate(df.iterrows(), start=5):
            bg = LGR if ri % 2 == 0 else WHT
            ch = float(row.get("change_24h", 0))

            def hc(ci, val, aln="right", fc="1C1C1C", fmt=None, bold=False):
                c = ws4.cell(ri, ci, val)
                c.font = Font(name="Arial", size=8, bold=bold, color=fc)
                c.alignment = (ctr() if aln == "center" else
                               (rgt() if aln == "right" else lft()))
                c.border = th(); c.fill = sf(bg)
                if fmt: c.number_format = fmt

            hc(1, str(row.get("timestamp", "")), "center", "666666")
            hc(2, str(row.get("rank", "")),      "center")
            hc(3, str(row.get("name", "")),      "left",  bold=True)
            hc(4, str(row.get("symbol", "")),    "center", BLU, bold=True)
            hc(5, float(row.get("price", 0)),    fmt="$#,##0.0000")
            change_cell(ws4, ri, 6, ch, bg)
            hc(7, float(row.get("market_cap",   0)), fmt="$#,##0")
            hc(8, float(row.get("volume_24h",   0)), fmt="$#,##0")
            ws4.row_dimensions[ri].height = 16
    else:
        ws4.cell(5, 1).value = "No history yet — run again to populate this sheet."

    # ── Save workbook ─────────────────────────────────────────────────────────
    path = CONFIG["excel_file"]
    wb.save(path)
    print(col("  ✔  Excel Report   → " + path, C.GREEN))


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 10 — TERMINAL DASHBOARD
#  These functions print the colorful ASCII output to the console.
# ══════════════════════════════════════════════════════════════════════════════

def save_text_report(coins, pnl, alerts, score):
    """
    Appends a plain-text summary of the session to session_report.txt.
    Useful for keeping a log you can read without opening Excel.
    """
    now   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total = sum(r["value_usd"] for r in pnl) if pnl else 0
    gainer = max(coins, key=lambda x: x["change_24h"]) if coins else None
    loser  = min(coins, key=lambda x: x["change_24h"]) if coins else None

    lines = [
        "=" * 68,
        "  CRYPTOPULSE  ·  SESSION REPORT  ·  " + now,
        "=" * 68,
        "  Coins scraped    : {}".format(len(coins)),
        "  Sentiment score  : {}/100".format(score),
        "  Portfolio value  : {}".format(
            "${:,.2f}".format(total) if pnl else "N/A"),
        "  Top gainer       : {} ({})".format(
            gainer["symbol"], fmt_change(gainer["change_24h"])) if gainer else "",
        "  Top loser        : {} ({})".format(
            loser["symbol"],  fmt_change(loser["change_24h"]))  if loser  else "",
        "  Alerts fired     : {}".format(len(alerts)),
        "",
        "  {:<4}  {:<10}  {:>14}  {:>10}  {:>14}  {:>7}".format(
            "#", "SYMBOL", "PRICE", "24H CHG", "MARKET CAP", "TIER"),
        "  " + "─" * 64,
    ]
    for c in coins:
        tier_lbl, _ = classify_tier(c["market_cap"])
        lines.append("  {:<4}  {:<10}  {:>14}  {:>10}  {:>14}  {:>7}".format(
            c["rank"], c["symbol"], fmt_price(c["price"]),
            fmt_change(c["change_24h"]), fmt_mcap(c["market_cap"]), tier_lbl
        ))
    lines.append("")

    path = CONFIG["report_file"]
    with open(path, "a", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(col("  ✔  Text Report    → " + path, C.GREEN))


def divider(w=80, char="─", color=C.DIM):
    print(col("  " + char * w, color))


def print_banner():
    print(col(
        "\n"
        "  ╔══════════════════════════════════════════════════════════════════════════╗\n"
        "  ║  ██████╗██████╗ ██╗   ██╗██████╗ ████████╗ ██████╗      █████╗ ██╗      ║\n"
        "  ║ ██╔════╝██╔══██╗╚██╗ ██╔╝██╔══██╗╚══██╔══╝██╔═══██╗    ██╔══██╗██║      ║\n"
        "  ║ ██║     ██████╔╝ ╚████╔╝ ██████╔╝   ██║   ██║   ██║    ███████║██║      ║\n"
        "  ║ ██║     ██╔══██╗  ╚██╔╝  ██╔═══╝    ██║   ██║   ██║    ██╔══██║██║      ║\n"
        "  ║ ╚██████╗██║  ██║   ██║   ██║        ██║   ╚██████╔╝    ██║  ██║███████╗ ║\n"
        "  ║  ╚═════╝╚═╝  ╚═╝   ╚═╝   ╚═╝        ╚═╝    ╚═════╝     ╚═╝  ╚═╝╚══════╝ ║\n"
        "  ║          Next-Generation Cryptocurrency Intelligence Platform             ║\n"
        "  ╚══════════════════════════════════════════════════════════════════════════╝",
        C.CYAN + C.BOLD
    ))


def print_market_table(coins):
    print(col(
        "\n  ┌────────────────────────────────────────────────────────────────────────┐",
        C.DIM))
    print(col(
        "  │              📊  LIVE MARKET  ·  CoinMarketCap Top Coins              │",
        C.WHITE + C.BOLD))
    print(col(
        "  └────────────────────────────────────────────────────────────────────────┘",
        C.DIM))
    print(col(
        "  {:<4}  {:<13}  {:<6}  {:<7}  {:>13}  {:>11}  {:>14}  {:>12}".format(
            "#", "NAME", "SYMBOL", "TIER", "PRICE (USD)",
            "24H CHANGE", "MARKET CAP", "VOLUME 24H"),
        C.YELLOW + C.BOLD))
    divider(84)
    for c in coins:
        ch      = c["change_24h"]
        ch_col  = C.GREEN if ch >= 0 else C.RED
        tl, tc  = classify_tier(c["market_cap"])
        print("  {:<4}  {:<13}  {:<6}  {}  {}  {}  {:>14}  {:>12}".format(
            col(str(c["rank"]),   C.DIM),
            col(c["name"][:12],  C.WHITE + C.BOLD),
            col(c["symbol"],     C.CYAN),
            col("{:<7}".format(tl),                     tc),
            col("{:>13}".format(fmt_price(c["price"])), C.WHITE),
            col("{:>11}".format(fmt_change(ch)),         ch_col),
            col(fmt_mcap(c["market_cap"]),               C.MAGENTA),
            col(fmt_vol(c["volume_24h"]),                C.DIM),
        ))
    divider(84)


def print_spark_chart(coins):
    """
    Visual bar chart in the terminal.
    Each █ block represents ~0.4% of price change.
    Green bars for gains, red bars for losses.
    """
    print(col("\n  ⚡  24H PRICE CHANGE  (spark bars — each █ ≈ 0.4%)", C.BLUE + C.BOLD))
    divider(64)
    for c in coins:
        ch  = c["change_24h"]
        bar = ("█" * min(int(abs(ch) * 2.5), 36)).ljust(36)
        sign = "+" if ch >= 0 else "-"
        line = "  {:<7}  {}{}%  {}".format(
            c["symbol"], sign, "{:.2f}".format(abs(ch)), bar)
        print(col(line, C.GREEN if ch >= 0 else C.RED))
    print()


def print_movers(coins):
    """Side-by-side panels showing top gainers and top losers."""
    gainers = sorted(coins, key=lambda x: x["change_24h"], reverse=True)[:CONFIG["top_gainers"]]
    losers  = sorted(coins, key=lambda x: x["change_24h"])[:CONFIG["top_losers"]]

    print(col("  ┌───────────────────────────────────┐   ┌───────────────────────────────────┐", C.DIM))
    print(
        col("  │   🏆  TOP GAINERS  (24H)           │", C.GREEN + C.BOLD) + "   " +
        col("│   🔻  TOP LOSERS   (24H)           │", C.RED + C.BOLD)
    )
    print(col("  ├───────────────────────────────────┤   ├───────────────────────────────────┤", C.DIM))
    for i in range(max(len(gainers), len(losers))):
        g_out = col("  │                                   │", C.DIM)
        l_out = col("   │                                   │", C.DIM)
        if i < len(gainers):
            g = gainers[i]
            g_out = (col("  │  ", C.DIM) +
                     col("{}. {:<6}  {}  {}".format(
                         i+1, g["symbol"], fmt_change(g["change_24h"]),
                         fmt_price(g["price"])), C.GREEN) +
                     col("  │", C.DIM))
        if i < len(losers):
            l = losers[i]
            l_out = (col("   │  ", C.DIM) +
                     col("{}. {:<6}  {}  {}".format(
                         i+1, l["symbol"], fmt_change(l["change_24h"]),
                         fmt_price(l["price"])), C.RED) +
                     col("  │", C.DIM))
        print(g_out + l_out)
    print(col("  └───────────────────────────────────┘   └───────────────────────────────────┘", C.DIM))
    print()


def print_sentiment(score, label):
    """Gauge bar showing market sentiment from 0 (extreme fear) to 100 (extreme greed)."""
    filled  = int(score / 100 * 32)
    bar     = ("█" * filled).ljust(32)
    bar_col = C.GREEN if score >= 60 else (C.YELLOW if score >= 40 else C.RED)
    print(col("  📡  MARKET SENTIMENT ENGINE", C.CYAN + C.BOLD))
    divider(52)
    print("  Score   :  " + col("{}/100".format(score), bar_col + C.BOLD))
    print("  Status  :  " + label)
    print("  Gauge   :  [" + col(bar, bar_col) + "]")
    print()


def print_dominance(dominance, coins):
    """Dominance bars — wider bar = larger share of combined market cap."""
    print(col("  🏆  MARKET DOMINANCE  (within Top-{})".format(len(coins)),
              C.MAGENTA + C.BOLD))
    divider(58)
    for sym, pct in sorted(dominance.items(), key=lambda x: -x[1]):
        bar = ("▓" * max(1, int(pct / 2))).ljust(30)
        dc  = C.CYAN if pct >= 20 else (C.YELLOW if pct >= 5 else C.DIM)
        print("  {:<7}  {:<30}  {}".format(
            col(sym, dc + C.BOLD),
            col(bar, dc),
            col("{:.1f}%".format(pct), dc),
        ))
    print()


def print_alerts(alerts):
    if alerts:
        print(col("  ╔══════════════════════════════════════════╗", C.YELLOW))
        print(col("  ║   ⚠️   LIVE MARKET ALERTS TRIGGERED     ║", C.YELLOW + C.BOLD))
        print(col("  ╚══════════════════════════════════════════╝", C.YELLOW))
        for a in alerts:
            print(a)
    else:
        print(col("  ✅  No major alerts  (surge >{:.0f}%  crash <{:.0f}%)".format(
            CONFIG["alert_surge"], abs(CONFIG["alert_crash"])), C.DIM))
    print()


def print_portfolio(pnl):
    if not pnl:
        print(col("  ℹ️  Portfolio empty — add your holdings to CONFIG['portfolio'].", C.DIM))
        return
    total = sum(r["value_usd"] for r in pnl)
    print(col("  ╔════════════════════════════════════════════════════════════╗", C.MAGENTA))
    print(col("  ║                💼  PORTFOLIO  P&L  SNAPSHOT               ║", C.MAGENTA + C.BOLD))
    print(col("  ╠════════════════════════════════════════════════════════════╣", C.MAGENTA))
    print(col("  ║  {:<10}  {:>10}  {:>14}  {:>13}  {:>5}  ║".format(
        "COIN", "QTY", "PRICE", "VALUE USD", "SHARE"), C.YELLOW + C.BOLD))
    print(col("  ╠════════════════════════════════════════════════════════════╣", C.MAGENTA))
    for r in pnl:
        share = r["value_usd"] / total * 100 if total > 0 else 0
        print("  ║  {:<10}  {:>10}  {:>14}  {:>13}  {:>4}%  ║".format(
            col(r["symbol"],                       C.CYAN),
            col("{:.4f}".format(r["qty"]),          C.WHITE),
            col(fmt_price(r["price"]),              C.WHITE),
            col("${:,.2f}".format(r["value_usd"]), C.GREEN),
            col("{:.1f}".format(share),             C.YELLOW),
        ))
    print(col("  ╠════════════════════════════════════════════════════════════╣", C.MAGENTA))
    print("  ║  {:<36}  {:>13}       ║".format(
        col("TOTAL PORTFOLIO VALUE", C.YELLOW + C.BOLD),
        col("${:,.2f}".format(total), C.GREEN + C.BOLD),
    ))
    print(col("  ╚════════════════════════════════════════════════════════════╝", C.MAGENTA))
    print()


def repair_history_csv(path):
    """
    One-time auto-repair: if the history CSV contains rows with the wrong
    number of columns (e.g. from old session-snapshot data being mixed in),
    this rewrites the file keeping only valid 8-column coin rows.
    Called automatically before reading — the user never needs to run this.
    """
    if not os.path.isfile(path):
        return
    fields   = ["timestamp","rank","name","symbol","price",
                "change_24h","market_cap","volume_24h"]
    good_rows = []
    header_written = False

    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue                        # skip blank / comment lines
            parts = line.split(",")
            if len(parts) == 8:
                if not header_written:
                    # First valid 8-column line might be the header
                    if parts[0].lower() == "timestamp":
                        good_rows.append(line)
                        header_written = True
                        continue
                good_rows.append(line)

    if not good_rows:
        return  # Nothing to save

    # If no header row survived, prepend one
    if good_rows[0].split(",")[0].lower() != "timestamp":
        good_rows.insert(0, ",".join(fields))

    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write("\n".join(good_rows) + "\n")


def print_historical():
    """Read the CSV history file and show session-over-session trends using pandas."""
    path = CONFIG["csv_history"]
    if not os.path.isfile(path):
        print(col("  ℹ️  No history yet — run again to build trend data.", C.DIM))
        return

    # Auto-repair the file before reading — fixes mixed/corrupt rows silently
    repair_history_csv(path)

    # Read CSV safely:
    #   comment="#"        → skip lines starting with # (metadata rows)
    #   on_bad_lines="skip"→ ignore any row with wrong column count
    #   dtype=str          → read all as strings first, convert below
    try:
        df = pd.read_csv(
            path,
            comment="#",
            on_bad_lines="skip",
            dtype=str,
            engine="python",
        )
    except Exception as e:
        print(col("  ⚠  Could not read history CSV: " + str(e), C.YELLOW))
        return

    # Keep only rows that have the exact 8 columns we expect
    expected = {"timestamp","rank","name","symbol","price","change_24h","market_cap","volume_24h"}
    if df.empty or not expected.issubset(set(df.columns)):
        print(col("  ⚠  History CSV has unexpected columns — skipping trend analysis.", C.YELLOW))
        return

    # Drop any row where symbol or price is missing / not a real coin row
    df = df[df["symbol"].notna() & df["price"].notna()]
    df = df[~df["symbol"].str.startswith("=")]   # Drop section-header rows
    df = df[~df["symbol"].str.startswith("#")]   # Drop comment rows

    # Convert price to numeric — coerce invalid strings to NaN and drop them
    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df = df.dropna(subset=["price"])
    df["change_24h"] = pd.to_numeric(df["change_24h"], errors="coerce").fillna(0)

    if df.empty or "symbol" not in df.columns:
        return

    print(col("  ┌─────────────────────────────────────────────────────────────────────┐", C.DIM))
    print(col("  │           📈  HISTORICAL TREND ENGINE  (all sessions)               │", C.CYAN + C.BOLD))
    print(col("  └─────────────────────────────────────────────────────────────────────┘", C.DIM))
    print(col("  {:<9}  {:>9}  {:>13}  {:>13}  {:>13}  {:>10}".format(
        "SYMBOL", "SESSIONS", "AVG PRICE", "MIN PRICE", "MAX PRICE", "TREND"),
        C.YELLOW + C.BOLD))
    divider(76)

    for sym in df["symbol"].unique():
        sub  = df[df["symbol"] == sym].copy().sort_values("timestamp")
        sess = len(sub)
        avg  = sub["price"].mean()
        mn   = sub["price"].min()
        mx   = sub["price"].max()
        pct  = ((sub.iloc[-1]["price"] - sub.iloc[0]["price"])
                / sub.iloc[0]["price"] * 100) if sub.iloc[0]["price"] > 0 else 0
        tc   = C.GREEN if pct >= 0 else C.RED
        print("  {:<9}  {:>9}  {:>13}  {:>13}  {:>13}  {}".format(
            col(sym,          C.CYAN),
            col(str(sess),    C.WHITE),
            col(fmt_price(avg), C.WHITE),
            col(fmt_price(mn),  C.RED),
            col(fmt_price(mx),  C.GREEN),
            col("{} {:.2f}%".format("▲" if pct >= 0 else "▼", abs(pct)), tc),
        ))
    divider(76)
    print()


def print_final_summary():
    """Final output summary box listing all 5 saved files."""
    print()
    print(col("  ╔══════════════════════════════════════════════════════════════════════╗", C.GREEN))
    print(col("  ║   ✅  CRYPTOPULSE SESSION COMPLETE  —  5 Files Saved               ║", C.GREEN + C.BOLD))
    print(col("  ╠══════════════════════════════════════════════════════════════════════╣", C.GREEN))
    print(col("  ║   📊  Excel Workbook  →  " + CONFIG["excel_file"].ljust(44)    + "║", C.GREEN))
    print(col("  ║   📁  History CSV     →  " + CONFIG["csv_history"].ljust(44)   + "║", C.GREEN))
    print(col("  ║   📋  Session CSV     →  " + CONFIG["csv_session"].ljust(44)   + "║", C.GREEN))
    print(col("  ║   💼  Portfolio CSV   →  " + CONFIG["csv_portfolio"].ljust(44) + "║", C.GREEN))
    print(col("  ║   📝  Text Report     →  " + CONFIG["report_file"].ljust(44)   + "║", C.GREEN))
    print(col("  ╚══════════════════════════════════════════════════════════════════════╝", C.GREEN))
    print()


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 11 — MAIN ENTRY POINT
#  This is the only code that runs when you type:  python crypto_tracker.py
#  It calls all the functions above in the correct order.
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # ── 1. Banner + session info ──────────────────────────────────────────────
    print_banner()
    print(col("  🕐  Started   : " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"), C.DIM))
    print(col("  ⚙️   Mode      : " + ("Headless (silent)" if CONFIG["headless"]
                                        else "Visible browser"), C.DIM))
    print(col("  📂  Output    : ./" + OUTPUT_DIR + "/", C.DIM))
    print()

    # ── 2. Launch Chrome and scrape CoinMarketCap ─────────────────────────────
    #    We use try/finally so the browser is always closed even if scraping fails.
    driver = create_driver()
    coins  = []
    try:
        coins = scrape(driver)
    finally:
        driver.quit()
        print(col("  ✔  Browser closed.\n", C.DIM))

    if not coins:
        print(col("  ✗  No data scraped. Check internet connection.", C.RED))
        sys.exit(1)

    print(col("  ✔  Scraped {} coins successfully.\n".format(len(coins)), C.GREEN))

    # ── 3. Compute derived intelligence from raw coin data ────────────────────
    display   = apply_filters(coins)          # Price-filtered subset for display
    pnl       = calc_portfolio(coins)         # Portfolio P&L rows
    dominance = compute_dominance(coins)      # Market share dict
    score, label = compute_sentiment(coins)   # Sentiment score + label
    alerts    = generate_alerts(display)      # Alert messages

    # ── 4. Print terminal dashboard ───────────────────────────────────────────
    print_market_table(display)
    print_spark_chart(display)
    print_movers(display)
    print_sentiment(score, label)
    print_dominance(dominance, coins)
    print_alerts(alerts)
    print_portfolio(pnl)

    # ── 5. Save all output files ──────────────────────────────────────────────
    print(col("  💾  Saving output files...\n", C.CYAN + C.BOLD))

    save_csv_history(coins)                           # FILE 1: history log (append)
    save_csv_session(coins, pnl, dominance, score)    # FILE 2: session snapshot (overwrite)
    save_csv_portfolio(pnl)                           # FILE 3: portfolio log (append)

    print_historical()                                # Trend analysis using pandas

    save_excel(coins, pnl, dominance, score, label)   # FILE 4: Excel workbook
    save_text_report(coins, pnl, alerts, score)       # FILE 5: text report

    # ── 6. Final summary ──────────────────────────────────────────────────────
    print_final_summary()


# Standard Python idiom: only run main() when this file is executed directly,
# not when it is imported as a module by another script.
if __name__ == "__main__":
    main()